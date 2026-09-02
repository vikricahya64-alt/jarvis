"""
E2B Code Interpreter executor: the "Hands" of J.A.R.V.I.S.

Securely runs Python/JS in an isolated cloud sandbox, captures stdout,
and downloads any generated files (PNG/CSV/JSON/PDF).
"""
import os
import base64

try:
    from e2b_code_interpreter import Sandbox
    E2B_AVAILABLE = True
except ImportError:
    E2B_AVAILABLE = False


def execute_code(code: str, language: str = "python") -> dict:
    """
    Execute code in an E2B sandbox and return stdout + generated files.

    Args:
        code: The source code to run.
        language: 'python' or 'javascript'.

    Returns:
        A dict: {success, stdout, stderr, files: [{name, data_b64, mime}]}
    """
    if not E2B_AVAILABLE:
        return {
            "success": False,
            "stdout": "",
            "stderr": "e2b-code-interpreter library not installed",
            "files": [],
        }

    api_key = os.getenv("E2B_API_KEY")
    if not api_key:
        return {
            "success": False,
            "stdout": "",
            "stderr": "E2B_API_KEY is not configured",
            "files": [],
        }

    return _run_sandbox(code, language)


def _run_sandbox(code: str, language: str) -> dict:
    sbx = None
    try:
        sbx = Sandbox()
        if language == "javascript":
            sbx.commands.run("npm install -g node 2>/dev/null || true")
            result = sbx.run_code(code, language="javascript")
        else:
            # Ensure common libs are available (cached between runs)
            sbx.commands.run(
                "pip install --quiet pandas matplotlib numpy 2>/dev/null || true"
            )
            result = sbx.run_code(code, language="python")

        stdout = result.text if result.text else ""
        stderr = getattr(result, "stderr", "") or ""

        # Download generated files referenced by relative paths
        files = _collect_files(sbx, ["png", "csv", "json", "pdf", "xlsx", "html"])

        return {
            "success": True,
            "stdout": stdout,
            "stderr": stderr,
            "files": files,
        }
    except Exception as exc:
        return {
            "success": False,
            "stdout": "",
            "stderr": f"E2B sandbox error: {exc}",
            "files": [],
        }
    finally:
        if sbx is not None:
            try:
                sbx.kill()
            except Exception:
                pass


def _collect_files(sbx, extensions):
    """
    Scan common working directories in the sandbox for generated files.
    E2B's run_code result exposes `files`; we also probe openai files.
    """
    files = []
    try:
        # Best effort: list files in /home/user and current dir
        for ext in extensions:
            listing = sbx.files.list(f"/home/user/*.{ext}")
            for handle in listing:
                data = sbx.files.read(f"/home/user/{handle.name}").encode("latin1")
                files.append({
                    "name": handle.name,
                    "data_b64": base64.b64encode(data).decode(),
                    "mime": _mime_for(ext),
                })
    except Exception:
        pass
    return files


def _mime_for(ext: str) -> str:
    mapping = {
        "png": "image/png",
        "csv": "text/csv",
        "json": "application/json",
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "html": "text/html",
        "txt": "text/plain",
    }
    return mapping.get(ext, "application/octet-stream")


def extract_document(fname: str, data_b64: str) -> dict:
    """
    Extract readable text from a PDF/DOCX/XLSX by running Python inside an
    E2B sandbox. The file is embedded as base64 in the script (avoids
    depending on the sandbox filesystem write API).

    Returns: {success, text, error}
    """
    if not E2B_AVAILABLE or not os.getenv("E2B_API_KEY"):
        return {"success": False, "text": "", "error": "E2B tidak tersedia"}

    ext = (fname.rsplit(".", 1)[-1] if "." in fname else "").lower()
    script = (
        "import base64, traceback\n"
        "try:\n"
        f"    data = base64.b64decode({data_b64!r})\n"
        "    open('/home/user/input', 'wb').write(data)\n"
        f"    ext = {ext!r}\n"
        "    text = ''\n"
        "    if ext == 'pdf':\n"
        "        from pypdf import PdfReader\n"
        "        r = PdfReader('/home/user/input')\n"
        "        text = '\\n'.join((p.extract_text() or '') for p in r.pages)\n"
        "    elif ext == 'docx':\n"
        "        from docx import Document\n"
        "        d = Document('/home/user/input')\n"
        "        text = '\\n'.join(p.text for p in d.paragraphs)\n"
        "    elif ext == 'xlsx':\n"
        "        from openpyxl import load_workbook\n"
        "        wb = load_workbook('/home/user/input', read_only=True, data_only=True)\n"
        "        out = []\n"
        "        for ws in wb.worksheets:\n"
        "            out.append('== Sheet: ' + ws.title)\n"
        "            for row in ws.iter_rows(values_only=True):\n"
        "                out.append(' | '.join('' if c is None else str(c) for c in row))\n"
        "        text = '\\n'.join(out)\n"
        "    text = text.strip()\n"
        "    open('/home/user/out.txt', 'w', encoding='utf-8').write(text)\n"
        "except Exception:\n"
        "    open('/home/user/err.txt', 'w').write(traceback.format_exc()[-500:])\n"
    )

    sbx = None
    try:
        sbx = Sandbox()
        sbx.commands.run(
            "pip install --quiet pypdf python-docx openpyxl 2>/dev/null || true"
        )
        result = sbx.run_code(script, language="python")
        out = None
        err_text = None
        try:
            if sbx.files.exists("/home/user/out.txt"):
                out = sbx.files.read("/home/user/out.txt").encode("latin1").decode("utf-8", errors="replace")
        except Exception:
            pass
        try:
            if sbx.files.exists("/home/user/err.txt"):
                err_text = sbx.files.read("/home/user/err.txt").encode("latin1").decode("utf-8", errors="replace")
        except Exception:
            pass
        if out is None or not out.strip():
            return {"success": False, "text": "",
                    "error": (err_text or "gagal ekstrak")[:300]}
        return {"success": True, "text": out, "error": ""}
    except Exception as exc:
        return {"success": False, "text": "", "error": str(exc)[:300]}
    finally:
        if sbx is not None:
            try:
                sbx.kill()
            except Exception:
                pass


